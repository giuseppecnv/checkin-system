import os
import psycopg2
from psycopg2.extras import DictCursor
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook, load_workbook

# Configurazione connessione (La stringa verrà presa da Vercel)
DATABASE_URL = os.getenv("DATABASE_URL")

# Se la stringa di Supabase inizia con postgres://, SQLAlchemy/Psycopg2 a volte
# preferiscono postgresql://. Questo piccolo fix previene errori:
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

def get_connection():
    """Crea una nuova connessione a Supabase."""
    return psycopg2.connect(DATABASE_URL)

def add_checkin(vdash: str):
    conn = get_connection()
    cur = conn.cursor()
    
    # Forza il minuscolo per combaciare con la tabella 'users'
    vdash_clean = vdash.lower().strip()
    
    now = datetime.now()
    checkin_time = now.strftime("%H:%M:%S")
    checkin_date = now.date().isoformat()

    try:
        cur.execute(
            "INSERT INTO checkins (vdash, checkin_time, checkin_date) VALUES (%s, %s, %s)",
            (vdash_clean, checkin_time, checkin_date)
        )
        conn.commit()
    except Exception as e:
        print(f"Errore add_checkin: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()

def add_checkout(vdash: str):
    conn = get_connection()
    cur = conn.cursor()

    now = datetime.now()
    checkout_time = now.time().isoformat(timespec='seconds')
    today = now.date().isoformat()

    cur.execute(
        """
        UPDATE checkins
        SET checkout_time = %s
        WHERE UPPER(vdash) = UPPER(%s)
        AND checkin_date = %s
        AND checkout_time IS NULL
        """,
        (checkout_time, vdash, today)
    )

    conn.commit()
    cur.close()
    conn.close()

def get_all_vdash():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT vdash FROM users;")
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return [row[0].upper() for row in rows]

def is_already_checked_in(vdash: str) -> bool:
    """Verifica se l'utente ha già effettuato il check-in oggi (case-insensitive)."""
    conn = get_connection()
    cur = conn.cursor()
    
    # Prendiamo la data di oggi in formato stringa ISO (YYYY-MM-DD)
    today = date.today().isoformat()
    
    # Puliamo l'input dell'utente
    vdash_clean = vdash.lower().strip()

    try:
        # Usiamo LOWER() nella query SQL per essere sicuri al 100% della corrispondenza
        cur.execute(
            "SELECT id FROM checkins WHERE LOWER(vdash) = %s AND checkin_date = %s;",
            (vdash_clean, today)
        )
        row = cur.fetchone()
        return row is not None
    except Exception as e:
        print(f"Errore is_already_checked_in: {e}")
        return False
    finally:
        cur.close()
        conn.close()

def is_already_checked_out(vdash: str) -> bool:
    conn = get_connection()
    cur = conn.cursor()

    today = date.today().isoformat()

    cur.execute(
        """
        SELECT id FROM checkins 
        WHERE UPPER(vdash) = UPPER(%s) 
        AND checkin_date = %s 
        AND checkout_time IS NOT NULL;
        """,
        (vdash, today)
    )

    exists = cur.fetchone() is not None
    cur.close()
    conn.close()
    return exists

def get_checkins_by_date(target_date: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            c.vdash,
            c.checkin_time,
            c.checkout_time,
            u.first_name,
            u.middle_name,
            u.last_name
        FROM checkins c
        JOIN users u ON UPPER(c.vdash) = UPPER(u.vdash)
        WHERE c.checkin_date = %s
        ORDER BY c.checkin_time;
    """, (target_date,))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    formatted_rows = []
    for row in rows:
        vdash = row[0].upper()
        # Gestione del tempo (trasformazione in stringa HH:MM)
        checkin_t = str(row[1])[:5] if row[1] else ""
        checkout_t = str(row[2])[:5] if row[2] else ""
        
        first = row[3].upper() if row[3] else ""
        middle = row[4].upper() if row[4] else ""
        last = row[5].upper() if row[5] else ""
        full_name = f"{first} {middle} {last}".strip()

        formatted_rows.append((vdash, checkin_t, checkout_t, full_name))

    return formatted_rows

def get_user_by_token(token: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT vdash, first_name, middle_name, last_name FROM users WHERE token = %s",
        (token,)
    )
    
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row

def get_checkin_time(vdash: str):
    conn = get_connection()
    cur = conn.cursor()
    today = date.today().isoformat()

    cur.execute(
        "SELECT checkin_time FROM checkins WHERE UPPER(vdash) = UPPER(%s) AND checkin_date = %s",
        (vdash, today)
    )

    row = cur.fetchone()
    cur.close()
    conn.close()
    
    return str(row[0])[:5] if row and row[0] else None

# Nota: Su Vercel, questo file Excel verrà creato in una cartella temporanea
def export_date_to_excel(target_date: str):
    rows = get_checkins_by_date(target_date)
    if not rows: return

    xlsx_path = "/tmp/checkins.xlsx" # Percorso temporaneo per Vercel
    
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if target_date in wb.sheetnames:
        del wb[target_date]

    ws = wb.create_sheet(title=target_date)
    ws.append(["V-DASH", "FULL NAME", "CHECK-IN TIME", "CHECK-OUT TIME"])

    for row in rows:
        ws.append([row[0], row[3], row[1], row[2]])

    wb.save(xlsx_path)
