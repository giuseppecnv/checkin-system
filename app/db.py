import os
import psycopg2
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook, load_workbook

# Configurazione variabili d'ambiente
DATABASE_URL = os.getenv("DATABASE_URL")

def get_connection():
    """Crea una nuova connessione a Supabase con supporto SSL obbligatorio."""
    url = DATABASE_URL
    if not url:
        raise ValueError("La variabile DATABASE_URL non è impostata!")
        
    # Fix per protocollo postgresql
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    
    # Aggiunta sslmode per sicurezza su cloud (Vercel -> Supabase)
    if "sslmode=" not in url:
        separator = "&" if "?" in url else "?"
        url += f"{separator}sslmode=require"
        
    return psycopg2.connect(url)

def add_checkin(vdash: str):
    conn = get_connection()
    cur = conn.cursor()
    
    now = datetime.now()
    checkin_time = now.strftime("%H:%M:%S")
    checkin_date = now.date().isoformat()

    try:
        cur.execute(
            """
            INSERT INTO checkins (vdash, checkin_time, checkin_date)
            VALUES (%s, %s, %s)
            """,
            (vdash.upper(), checkin_time, checkin_date)
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Errore add_checkin: {e}")
    finally:
        cur.close()
        conn.close()

def add_checkout(vdash: str):
    conn = get_connection()
    cur = conn.cursor()

    now = datetime.now()
    checkout_time = now.strftime("%H:%M:%S")
    today = now.date().isoformat()

    try:
        cur.execute(
            """
            UPDATE checkins
            SET checkout_time = %s
            WHERE UPPER(vdash) = UPPER(%s)
            AND checkin_date = %s
            AND checkout_time IS NULL
            """,
            (checkout_time, vdash, today)
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Errore add_checkout: {e}")
    finally:
        cur.close()
        conn.close()

def get_all_vdash():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT vdash FROM users;")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [row[0].upper() for row in rows]

def is_already_checked_in(vdash: str) -> bool:
    conn = get_connection()
    cur = conn.cursor()
    today = date.today().isoformat()
    cur.execute(
        "SELECT id FROM checkins WHERE UPPER(vdash) = UPPER(%s) AND checkin_date = %s;",
        (vdash, today)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row is not None

def is_already_checked_out(vdash: str) -> bool:
    conn = get_connection()
    cur = conn.cursor()
    today = date.today().isoformat()
    cur.execute(
        """
        SELECT id FROM checkins 
        WHERE UPPER(vdash) = UPPER(%s) 
        AND checkin_date = %s 
        AND checkout_time IS NOT NULL;
        """,
        (vdash, today)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row is not None

def get_checkins_by_date(target_date: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            c.vdash,
            c.checkin_time,
            c.checkout_time,
            u.first_name,
            u.middle_name,
            u.last_name
        FROM checkins c
        JOIN users u ON UPPER(c.vdash) = UPPER(u.vdash)
        WHERE c.checkin_date = %s
        ORDER BY c.checkin_time;
    """, (target_date,))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    formatted_rows = []
    for row in rows:
        vdash = row[0].upper()
        checkin_t = str(row[1])[:5] if row[1] else ""
        checkout_t = str(row[2])[:5] if row[2] else ""
        first = row[3].upper() if row[3] else ""
        middle = row[4].upper() if row[4] else ""
        last = row[5].upper() if row[5] else ""
        full_name = f"{first} {middle} {last}".strip()
        formatted_rows.append((vdash, checkin_t, checkout_t, full_name))
    return formatted_rows

def get_checkout_time(vdash: str):
    conn = get_connection()
    cur = conn.cursor()
    today = date.today().isoformat()
    cur.execute(
        "SELECT checkout_time FROM checkins WHERE UPPER(vdash) = UPPER(%s) AND checkin_date = %s",
        (vdash, today)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return str(row[0])[:5] if row and row[0] else None

def export_date_to_excel(target_date: str):
    rows = get_checkins_by_date(target_date)
    if not rows: return
    
    xlsx_path = "/tmp/checkins.xlsx"
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if target_date in wb.sheetnames:
        del wb[target_date]

    ws = wb.create_sheet(title=target_date)
    ws.append(["V-DASH", "FULL NAME", "CHECK-IN TIME", "CHECK-OUT TIME"])
    for row in rows:
        ws.append([row[0], row[3], row[1], row[2]])
    wb.save(xlsx_path)

def get_user_by_token(token: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT vdash, first_name, middle_name, last_name FROM users WHERE token = %s",
        (token,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row

def get_checkin_time(vdash: str):
    conn = get_connection()
    cur = conn.cursor()
    today = date.today().isoformat()
    cur.execute(
        "SELECT checkin_time FROM checkins WHERE UPPER(vdash) = UPPER(%s) AND checkin_date = %s",
        (vdash, today)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    return str(row[0])[:5] if row and row[0] else None
